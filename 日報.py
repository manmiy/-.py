import streamlit as st
import os
import pandas as pd
from io import StringIO, BytesIO
from PIL import Image
from datetime import timedelta, date, datetime
import re
from openpyxl.utils import get_column_letter
from google import genai
from google.oauth2 import service_account
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import jpholiday
from concurrent.futures import ThreadPoolExecutor, as_completed 

# ==========================================
# 1. 初期設定と共通関数
# ==========================================
st.set_page_config(page_title="日報 (Vertex AI)", page_icon="📝", layout="wide")

# 祝日判定ツール
def is_holiday_jp(d):
    return d.weekday() >= 5 or jpholiday.is_holiday(d)

if 'extracted_df' not in st.session_state:
    st.session_state.extracted_df = None

if 'is_reading' not in st.session_state:
    st.session_state.is_reading = False

def clear_extracted_data():
    st.session_state.extracted_df = None
    st.session_state.is_reading = False

# 時刻ヘルパー関数
def time_to_min(t_str):
    try:
        parts = t_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0

def min_to_str(m):
    return f"{m//60}:{m%60:02d}"

# ==========================================
# 2. Excel生成ロジック (はみ出た時間外の自動分割版)
# ==========================================
def create_filled_excel(df_extracted, sheet1_name="人工集計", sheet2_name="日報明細", target_year=2024, reg_start_time="8:00", reg_end_time="17:00"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet1_name
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    reg_start_min = time_to_min(reg_start_time) # 8:00 -> 480分
    reg_end_min = time_to_min(reg_end_time)     # 17:00 -> 1020分

    raw_data_list = []
    unique_codes = []
    code_to_name = {}
    last_seen_code = ""

    # --- 1. 生データの抽出 ---
    if df_extracted is not None and not df_extracted.empty:
        for idx, row in df_extracted.iterrows():
            try:
                m_str = str(row.get('月', '')).replace('月', '').strip()
                d_str = str(row.get('日', '')).replace('日', '').strip()
                t_val = str(row.get('時間', '')).strip()

                code_val = str(row.get('工事コード', '')).strip()
                code_val = re.sub(r'[^a-zA-Z0-9-]', '', code_val)

                if not code_val or code_val.lower() in ['nan', 'none']:
                    code_val = last_seen_code
                else:
                    last_seen_code = code_val

                if code_val:
                    code_to_name[code_val] = ""

                content_val = str(row.get('業務内容', '')).strip()
                t_val = re.sub(r'[^0-9:]', '', t_val)
                if not t_val: continue

                m = int(float(m_str))
                d = int(float(d_str))

                if code_val and code_val not in unique_codes:
                    unique_codes.append(code_val)

                break_raw = row.get('休憩', None)
                if pd.isna(break_raw) or str(break_raw).strip() == "" or break_raw is None:
                    break_val = ""
                else:
                    try: break_val = float(break_raw)
                    except (ValueError, TypeError): break_val = ""

                raw_data_list.append({
                    'orig_idx': idx,
                    'm': m, 'd': d, 'break': break_val,
                    'time': t_val, 'code': code_val, 'content': content_val,
                    'site': str(row.get('現場名', ''))
                })
            except Exception as e:
                print(f"Excel前処理エラー: {e}")

    # --- 2. 日毎に連続する時間を追跡し、定時内外（8:00〜17:00）で自動分割 ---
    grouped_raw = {}
    for entry in raw_data_list:
        key = (entry['m'], entry['d'])
        if key not in grouped_raw: grouped_raw[key] = []
        grouped_raw[key].append(entry)

    data_list = []

    for (m, d), entries in grouped_raw.items():
        try:
            d_obj = date(target_year, m, d)
            is_hol = is_holiday_jp(d_obj)
        except:
            is_hol = False

        first_end_min = time_to_min(entries[0]['time'])
        # 朝の開始時間判定: 最初の終了時刻が8:00より前ならその時刻を開始地点にする
        curr_start = first_end_min if first_end_min < reg_start_min else reg_start_min

        for entry in entries:
            end_min = time_to_min(entry['time'])
            if end_min <= curr_start:
                curr_start = end_min
                continue

            s_min = curr_start
            e_min = end_min
            curr_start = end_min # 次の作業の始業時刻は現在の終業時刻

            # 休日の場合は時間帯にかかわらず全て「残業・休日出勤」枠へ配置
            if is_hol:
                ne = dict(entry)
                ne['start_time'] = min_to_str(s_min)
                ne['end_time'] = min_to_str(e_min)
                ne['category'] = "残業/時間外"
                data_list.append(ne)
                continue

            # A. 早朝残業 (0:00 〜 8:00)
            if s_min < reg_start_min:
                em_e = min(e_min, reg_start_min)
                ne = dict(entry)
                ne['start_time'] = min_to_str(s_min)
                ne['end_time'] = min_to_str(em_e)
                ne['category'] = "残業/時間外"
                ne['break'] = "" # 休憩は定時枠に優先割り当て
                data_list.append(ne)

            # B. 定時日勤 (8:00 〜 17:00)
            reg_s = max(s_min, reg_start_min)
            reg_e = min(e_min, reg_end_min)
            if reg_s < reg_e:
                ne = dict(entry)
                ne['start_time'] = min_to_str(reg_s)
                ne['end_time'] = min_to_str(reg_e)
                ne['category'] = "日勤"
                data_list.append(ne)

            # C. 夕方残業 (17:00 〜 24:00)
            if e_min > reg_end_min:
                ev_s = max(s_min, reg_end_min)
                ne = dict(entry)
                ne['start_time'] = min_to_str(ev_s)
                ne['end_time'] = min_to_str(e_min)
                ne['category'] = "残業/時間外"
                ne['break'] = "" # 休憩は定時枠に優先割り当て
                data_list.append(ne)

    # --- 3. ヘッダー構造の設定 ---
    headers_left = [
        ("B", "月"), ("C", "日"), ("D", "曜日"), ("E", "休・出"), ("F", "摘要"),
        ("G", "始業"), ("H", "終業"), ("I", "時間"), ("J", "休憩"), ("K", "時間"), ("L", "人工")
    ]

    def write_headers(ws, start_row, header_blocks):
        for col, name in headers_left:
            cell = ws[f"{col}{start_row}"]
            cell.value = name; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.font = Font(size=9, bold=True); cell.border = thin_border

        ws.merge_cells(f"K{start_row-2}:L{start_row-2}")
        for col in ("K", "L"): ws[f"{col}{start_row-2}"].border = thin_border
        c1 = ws[f"K{start_row-2}"]; c1.value = "実労"; c1.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"K{start_row-1}:L{start_row-1}")
        for col in ("K", "L"): ws[f"{col}{start_row-1}"].border = thin_border
        c2 = ws[f"K{start_row-1}"]; c2.value = "計"; c2.alignment = Alignment(horizontal="center", vertical="center")

        for block in header_blocks[1:]:
            h_code, h_name, col1, col2 = block
            ws.merge_cells(f"{col1}{start_row-2}:{col2}{start_row-2}")
            for col in (col1, col2): ws[f"{col}{start_row-2}"].border = thin_border
            cn = ws[f"{col1}{start_row-2}"]; cn.value = h_name; cn.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f"{col1}{start_row-1}:{col2}{start_row-1}")
            for col in (col1, col2): ws[f"{col}{start_row-1}"].border = thin_border
            cc = ws[f"{col1}{start_row-1}"]; cc.value = h_code; cc.alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{col1}{start_row}"] = "時間"; ws[f"{col2}{start_row}"] = "人工"
        for col in (col1, col2):
            cell = ws[f"{col}{start_row}"]
            cell.border = thin_border; cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    widths = {'A':4, 'B':4, 'C':4, 'D':4, 'E':6, 'F':12, 'G':6, 'H':6, 'I':8, 'J':6, 'K':10, 'L':9}
    for col, w in widths.items(): ws.column_dimensions[col].width = w

    header_blocks = [("計", "実労", "K", "L")]
    start_col_idx = 13
    for code in unique_codes:
        col1 = get_column_letter(start_col_idx); col2 = get_column_letter(start_col_idx + 1)
        header_blocks.append((code, code_to_name.get(code, ""), col1, col2))
        start_col_idx += 2
    header_blocks.append(("その他", "", get_column_letter(start_col_idx), get_column_letter(start_col_idx + 1)))
    start_col_idx += 2

    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    reg_data = [d for d in data_list if d.get('category') == "日勤"]
    hol_data = [d for d in data_list if d.get('category') != "日勤"]

    # --- 4. 行書き込み処理 ---
    def fill_rows(ws, data_rows, start_row_idx, total_label="計", is_continuous=False):
        r_idx = start_row_idx
        grouped = {}
        for entry in data_rows:
            md = (entry['m'], entry['d'])
            if md not in grouped: grouped[md] = []
            grouped[md].append(entry)

        def write_single_entry(r_idx, entry, curr_date):
            is_first = entry.get('is_first', True)
            ws[f'B{r_idx}'] = curr_date.month if is_first else ""
            ws[f'C{r_idx}'] = curr_date.day if is_first else ""
            ws[f'D{r_idx}'] = weekdays_jp[curr_date.weekday()] if is_first else ""
            ws[f'E{r_idx}'] = "〇" if is_first else ""
            ws[f'F{r_idx}'] = ""

            entry['final_excel_row'] = r_idx
            ws[f'G{r_idx}'] = entry['start_time']
            ws[f'H{r_idx}'] = entry['end_time']
            ws[f'I{r_idx}'] = f'=IF(OR(G{r_idx}="", H{r_idx}=""), "", (H{r_idx}-G{r_idx})*24)'
            b_val = entry.get('break')
            ws[f'J{r_idx}'] = None if b_val == "" else b_val
            ws[f'K{r_idx}'] = f'=IF(I{r_idx}="","",I{r_idx}-IF(J{r_idx}="",0,J{r_idx}))'
            ws[f'L{r_idx}'] = f'=IF(K{r_idx}="","",K{r_idx}/7)'

            col_t_idx = 13 + unique_codes.index(entry['code']) * 2 if entry.get('code') in unique_codes else start_col_idx - 2
            ws[f'{get_column_letter(col_t_idx)}{r_idx}'] = f'=K{r_idx}'
            ws[f'{get_column_letter(col_t_idx+1)}{r_idx}'] = f'=L{r_idx}'

            for col_idx in range(2, start_col_idx):
                cell = ws[f'{get_column_letter(col_idx)}{r_idx}']
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx in [7, 8]: cell.number_format = '[h]:mm'
                elif col_idx in [9, 10, 11]: cell.number_format = '0.00'
                elif col_idx == 12: cell.number_format = '0.0000'
                elif col_idx >= 13: cell.number_format = '0.00' if (col_idx % 2 != 0) else '0.0000'

        if is_continuous:
            if data_rows:
                sample = data_rows[0]
                sd = date(target_year, sample['m'], sample['d'])
            else:
                sd = date.today()

            if sd.day >= 21:
                period_start = date(sd.year, sd.month, 21)
            else:
                prev_m = 12 if sd.month == 1 else sd.month - 1
                prev_y = sd.year - 1 if sd.month == 1 else sd.year
                period_start = date(prev_y, prev_m, 21)

            next_m = period_start.month % 12 + 1
            next_y = period_start.year + (1 if period_start.month == 12 else 0)
            period_end = date(next_y, next_m, 20)

            curr = period_start
            while curr <= period_end:
                md = (curr.month, curr.day)
                entries = grouped.get(md, [])

                if not entries:
                    ws[f'B{r_idx}'] = curr.month
                    ws[f'C{r_idx}'] = curr.day
                    ws[f'D{r_idx}'] = weekdays_jp[curr.weekday()]
                    ws[f'E{r_idx}'] = "×"
                    ws[f'F{r_idx}'] = ""
                    for col_idx in range(7, start_col_idx):
                        ws[f'{get_column_letter(col_idx)}{r_idx}'] = ""
                    for col_idx in range(2, start_col_idx):
                        cell = ws[f'{get_column_letter(col_idx)}{r_idx}']
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    r_idx += 1
                else:
                    for t_idx, entry in enumerate(entries):
                        entry['is_first'] = (t_idx == 0)
                        write_single_entry(r_idx, entry, curr)
                        r_idx += 1
                curr += timedelta(days=1)
        else:
            for md, entries in grouped.items():
                for t_idx, entry in enumerate(entries):
                    entry['is_first'] = (t_idx == 0)
                    try: curr_date = date(target_year, entry['m'], entry['d'])
                    except: curr_date = date(target_year, 1, 1)
                    write_single_entry(r_idx, entry, curr_date)
                    r_idx += 1

        if r_idx > start_row_idx:
            ws.merge_cells(f"B{r_idx}:D{r_idx}"); ws[f"B{r_idx}"] = total_label
            ws[f"B{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"E{r_idx}"] = f'=COUNTIF(E{start_row_idx}:E{r_idx-1}, "〇")'
            ws[f"E{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx in range(9, start_col_idx):
                col_let = get_column_letter(col_idx)
                ws[f"{col_let}{r_idx}"] = f"=SUM({col_let}{start_row_idx}:{col_let}{r_idx-1})"
                if col_idx == 12 or (col_idx >= 14 and col_idx % 2 == 0):
                    ws[f"{col_let}{r_idx}"].number_format = '0.0000'
                else:
                    ws[f"{col_let}{r_idx}"].number_format = '0.00'
            for col_idx in range(2, start_col_idx):
                cell = ws[f'{get_column_letter(col_idx)}{r_idx}']; cell.border = thin_border; cell.font = Font(bold=True)
            r_idx += 1
        return r_idx

    # --- 5. 出力実行 ---
    write_headers(ws, 4, header_blocks)
    next_row = fill_rows(ws, reg_data, 5, total_label="日勤計", is_continuous=True)
    
    if hol_data:
        next_row += 3
        ws.merge_cells(f"A{next_row-1}:A{next_row + len(hol_data) + (1 if len(hol_data)>0 else 0)}")
        cell_a = ws[f"A{next_row-1}"]; cell_a.value = "残業・休日出勤"; cell_a.alignment = Alignment(horizontal="center", vertical="center", textRotation=255); cell_a.border = thin_border
        write_headers(ws, next_row, header_blocks)
        next_row = fill_rows(ws, hol_data, next_row + 1, total_label="残業計", is_continuous=False)

    sheet1_row_mapping = {}
    for idx_key, d in enumerate(data_list):
        if 'final_excel_row' in d:
            sheet1_row_mapping[idx_key] = d['final_excel_row']

    # --- 6. 2枚目シート(日報明細)の出力 ---
    ws2 = wb.create_sheet(title=sheet2_name)
    headers_s2 = ["月", "日", "始業", "終業", "時間", "休憩", "実労時間", "工事コード", "現場名", "業務内容"]
    for col_idx, h in enumerate(headers_s2, start=1):
        c = ws2.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True); c.border = thin_border; c.alignment = Alignment(horizontal="center", vertical="center")

    if data_list:
        r_idx = 3
        prev_m, prev_d = None, None
        
        for idx_key, entry in enumerate(data_list):
            m_num = entry['m']
            d_num = entry['d']

            if m_num == prev_m and d_num == prev_d:
                out_m, out_d = "", ""
            else:
                out_m, out_d = m_num, d_num
                prev_m, prev_d = m_num, d_num

            ws2[f'A{r_idx}'] = out_m
            ws2[f'B{r_idx}'] = out_d
            ws2[f'C{r_idx}'] = entry['start_time']
            ws2[f'D{r_idx}'] = entry['end_time']
            ws2[f'E{r_idx}'] = f'=IF(OR(C{r_idx}="", D{r_idx}=""), "", (D{r_idx}-C{r_idx})*24)'
            
            if idx_key in sheet1_row_mapping:
                s1_r = sheet1_row_mapping[idx_key]
                ws2[f'F{r_idx}'] = f'=IF(\'{sheet1_name}\'!J{s1_r}="","",\'{sheet1_name}\'!J{s1_r})'
            else:
                ws2[f'F{r_idx}'] = entry.get('break', '')

            ws2[f'G{r_idx}'] = f'=IF(E{r_idx}="","",E{r_idx}-IF(F{r_idx}="",0,F{r_idx}))'
            ws2[f'H{r_idx}'] = entry['code']
            ws2[f'I{r_idx}'] = entry.get('site', '')
            ws2[f'J{r_idx}'] = entry['content']

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                c = ws2[f'{col_letter}{r_idx}']
                c.border = thin_border
                if col_letter in ['I', 'J']:
                    c.alignment = Alignment(wrapText=True, vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

                if col_letter in ['C', 'D']: c.number_format = '[h]:mm'
                elif col_letter in ['E', 'F', 'G']: c.number_format = '0.00'

            r_idx += 1

        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 6
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 10
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 20
        ws2.column_dimensions['J'].width = 40

    return wb

# ==========================================
# 3. Streamlit UI
# ==========================================
st.title("📋 日報 ")

st.markdown("### ⚙️ 出力・対象年設定")
col1, col2, col3, col4 = st.columns(4)
with col1: target_year_val = st.number_input("📅 対象年", value=date.today().year, step=1)
with col2: filename_input = st.text_input("💾 保存ファイル名", value="日報データ.xlsx")
with col3: sheet1_input = st.text_input("📄 1枚目シート", value="人工集計")
with col4: sheet2_input = st.text_input("📄 2枚目シート", value="日報明細")

uploaded_files = st.file_uploader("ノートの画像をすべて選択してください", type=["jpg", "jpeg", "png","tif","tiff"], accept_multiple_files=True, on_change=clear_extracted_data)

if uploaded_files:
    if st.session_state.extracted_df is None:
        cols = st.columns(min(3, len(uploaded_files)) if len(uploaded_files) > 0 else 1)
        for i, file in enumerate(uploaded_files):
            with cols[i % 3]: st.image(Image.open(file), caption=file.name, use_container_width=True)
        if st.button("✨ 画像から読み取る！", type="primary", use_container_width=True, disabled=st.session_state.is_reading):
            st.session_state.is_reading = True; st.rerun()

    if st.session_state.is_reading:
        progress_bar = st.progress(0); status_text = st.empty(); all_dfs = []
        try:
            credentials_info = st.secrets["gcp_service_account"]
            creds = service_account.Credentials.from_service_account_info(credentials_info, scopes=["https://www.googleapis.com/auth/cloud-platform"])
            client = genai.Client(vertexai=True, project=credentials_info["project_id"], location="us-central1", credentials=creds)

            prompt_text = """
あなたは優秀なデータ入力アシスタントです。提供された画像は手書きの日報ノートです。
画像内の表は左から順に以下の列構成になっています：
1. 日付 (月/日)
2. 工事コード (1397-00 など)
3. 時間 (10:00 など)
4. 人工 (0.5, 2 など - これは抽出しますがExcel出力用には使いません)
5. 業務内容 (ML-750打設準備 など)

【抽出ルール】
- 「月」「日」「工事コード」「現場名」「時間」「業務内容」の6項目を抽出してください。
- 画像の最上部にある「〇年〇月分」などのタイトル表記や、表の枠外にあるメモ書きは絶対にデータとして抽出しないでください。罫線で囲まれたメインの表の中身のみを対象とします。
- ⚠️ 業務内容に該当する欄に、作業内容と一緒に「現場名（場所の名前など）」が書き込まれている場合があります。現場名は「現場名」として独立させ、残った純粋な作業内容を「業務内容」として別々に抽出してください。現場名がない場合は空欄としてください。
- ⚠️ 工事コードの「〃」といった同上を示す記号が書かれている場合は、記号をそのまま出力せず、直上の行から同じ内容を補完して出力してください。
- ⚠️ 工事コードが完全に空欄の場合も、直前の行と同じ工事コードを補完して出力してください。
- 時間に「-」や「〜」などの記号が含まれている場合、それらを除外して数字とコロンのみ（例: 10:30）にしてください。
- 業務内容や現場名に含まれる改行はスペースに置き換えてください。
- 出力は必ず「|」（パイプ記号）で区切ったテキスト形式（挨拶不要）でお願いします。

⚠️【最重要ルール：行の分割】
ノート上で「同じ日付・同じ工事コード」の枠内に、時間が複数行に分かれて書かれている場合（例：15:00、16:00...）、絶対にカンマ等で1つのセルにまとめないでください。
必ず「月」「日」「工事コード」「現場名」を上の行から補って、1つの時間につき1行（独立したデータ）として出力してください。

■ 出力フォーマット
1行目: 月|日|工事コード|現場名|時間|業務内容
2行目以降: | 区切りのデータ（計6列）
"""
            def process_image(file):
                try:
                    response = client.models.generate_content(model="gemini-2.5-flash", contents=[Image.open(file), prompt_text])
                    raw_text = response.text.strip()
                    
                    match = re.search(r'`{3}(?:.*?)\n(.*?)`{3}', raw_text, re.DOTALL)
                    base_text = match.group(1).strip() if match else raw_text

                    clean_lines = []
                    for line in base_text.split('\n'):
                        line = line.strip()
                        if not line or re.match(r'^\|?[\s\-]+\|?[\s\-\|]*$', line): continue
                        if line.startswith('|'): line = line[1:]
                        if line.endswith('|'): line = line[:-1]
                        clean_lines.append(line.strip())

                    col_names = ["月", "日", "工事コード", "現場名", "時間", "業務内容"]
                    df = pd.read_csv(StringIO('\n'.join(clean_lines)), sep='|', names=col_names, header=None, on_bad_lines='skip')

                    if not df.empty and str(df.iloc[0]['月']).strip() == "月":
                        df = df.iloc[1:].reset_index(drop=True)

                    df = df.dropna(subset=['月', '日', '業務内容'])

                    if not df.empty:
                        if '工事コード' in df.columns:
                            df['工事コード'] = df['工事コード'].astype(str).str.replace(r'[^a-zA-Z0-9-]', '', regex=True)
                        if '時間' in df.columns:
                            df['時間'] = df['時間'].astype(str).str.replace(r'[^0-9:]', '', regex=True)
                        
                        df['休憩'] = None
                        return df
                except Exception as e:
                    return {"error": str(e), "file_name": file.name, "raw_text": raw_text if 'raw_text' in locals() else ""}
                return None

            status_text.text(f"並列処理中... 全 {len(uploaded_files)} 枚")
            
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = [executor.submit(process_image, f) for f in uploaded_files]
                
                for idx, future in enumerate(as_completed(futures)):
                    result = future.result()
                    
                    if isinstance(result, pd.DataFrame):
                        all_dfs.append(result)
                    elif isinstance(result, dict) and "error" in result:
                        st.error(f"⚠️ {result['file_name']} の読み取りに失敗しました。（エラー: {result['error']}）")
                        with st.expander("AIの生データを確認"): st.text(result.get("raw_text", ""))
                    
                    progress_bar.progress((idx + 1) / len(uploaded_files))

            if all_dfs:
                st.session_state.extracted_df = pd.concat(all_dfs, ignore_index=True); st.session_state.is_reading = False; st.rerun()
            else: 
                st.session_state.is_reading = False; st.error("抽出失敗")
        except Exception as e: 
            st.session_state.is_reading = False; st.error(f"システムエラー: {e}")
    else:
        col1, col2 = st.columns(2)
        with col1:
            for f in uploaded_files: st.image(Image.open(f), use_container_width=True)
        with col2:
            st.subheader("📝 読み取り結果")
            edited_df = st.data_editor(st.session_state.extracted_df, num_rows="dynamic", use_container_width=True, height=850)
            if st.button("リセット", use_container_width=True): st.session_state.extracted_df = None; st.rerun()
            st.write("---")
            wb = create_filled_excel(edited_df, sheet1_input, sheet2_input, target_year_val)
            output = BytesIO(); wb.save(output); st.download_button("📥 ダウンロード", output.getvalue(), file_name=filename_input, use_container_width=True, type="primary")
